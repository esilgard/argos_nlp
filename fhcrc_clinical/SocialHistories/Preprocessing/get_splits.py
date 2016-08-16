import SystemUtilities.Configuration as c
import SystemUtilities.Globals as g
import os
import openpyxl
import collections


def load_caisis_silver_annotations():
    # Read in metadata file (excel), creating:
    #  dict of {doc_id:patient_id}
    #  dict of {patient_id : [gold labels] }
    wb = openpyxl.load_workbook(c.CAISIS_DIR)
    sheets = wb.get_sheet_names()
    mrn_caisis_dict = dict()
    caisis_gold_dict = dict()
    sheet_name = sheets[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    for i in range(1, g.CAISIS_ROWS, 1):  # the num of rows in metadata file
        caisis_id = str(sheet.cell(row=i, column=2).value)
        mrn = str(sheet.cell(row=i, column=1).value)
        mrn_caisis_dict[mrn] = caisis_id
        if caisis_id not in caisis_gold_dict.keys():
            caisis_gold_dict[caisis_id] = list()
        tobac_label = sheet.cell(row=i, column=3).value
        alc_label = sheet.cell(row=i, column=8).value
        caisis_gold_dict[caisis_id].append(alc_label)
        caisis_gold_dict[caisis_id].append(tobac_label)
    return mrn_caisis_dict, caisis_gold_dict


def write_new_splits_to_file(document_or_patient, dev_dict, test_dict, train_dict):
    tmp = 0
    if document_or_patient == "patient":
        od = collections.OrderedDict(sorted(dev_dict.items()))
        with open(c.patients_dev_gold_dir) as file:
            for key, value in od:
                file.write(key + "\t" + value)

    pass


def split_patient_data(patients_dev_dict, patients_test_dict, patients_train_dict, caisis_silver_dict):
    # Read in expanded data set and determine which documents/patients are not already accounted for in Flor's split
    full_data_set_dir = c.data_repo_dir
    unassigned = set()
    seen_patients = set()
    patient_level_split_assignments = dict()

    for file in os.listdir(full_data_set_dir):
        p_d = file.split("_")
        patient_id = p_d[0]
        doc_id = p_d[1]

        belongs = False
        if patient_id in patients_dev_dict:
            if patient_id not in seen_patients:
                patients_dev_dict[patient_id] += ",ALC=" + caisis_silver_dict[patient_id][
                    0]  # append silver alc label to existing gold
            belongs = True
            patient_level_split_assignments[patient_id] = "DEV"
        elif patient_id in patients_test_dict:
            if patient_id not in seen_patients:
                patients_test_dict[patient_id] += ",ALC=" + caisis_silver_dict[patient_id][
                    0]  # append silver alc label to existing gold
            belongs = True
            patient_level_split_assignments[patient_id] = "TEST"
        elif patient_id in patients_train_dict:
            if patient_id not in seen_patients:
                patients_train_dict[patient_id] += ",ALC=" + caisis_silver_dict[patient_id][
                    0]  # append silver alc label to existing gold
            belongs = True
            patient_level_split_assignments[patient_id] = "TRAIN"
        elif not belongs:  # put the patient in the unassigned bin
            unassigned.add(patient_id)
        seen_patients.add(patient_id)

    # iterate over the set of unassigned patients and split them up randomly into dev/test/train
    DEV = 1
    TEST = 2
    TRAIN = 5
    count = 0
    for patient in unassigned:
        try:
            if count == DEV:
                patients_dev_dict[patient] = caisis_silver_dict[patient][1] + ",ALC=" + caisis_silver_dict[patient][0]
                patient_level_split_assignments[patient] = "DEV"
            elif count == TEST:
                patients_test_dict[patient] = caisis_silver_dict[patient][1] + ",ALC=" + caisis_silver_dict[patient][0]
                patient_level_split_assignments[patient] = "TEST"
            else:
                patients_train_dict[patient] = caisis_silver_dict[patient][1] + ",ALC=" + caisis_silver_dict[patient][0]
                patient_level_split_assignments[patient] = "TRAIN"
            count += 1
            count = (count % 5)
        except KeyError:
            print "Error: Key " + patient + " does not have GOLD or SILVER labels! Adding patient to training data with unknown labels"
            patients_train_dict[patient] = "UNKNOWN" + ",ALC=UNKNOWN"

    # write new patient-level dev/test/train gold+silver alc+tob splits to file
    write_new_splits_to_file("patients", patients_dev_dict, patients_test_dict, patients_train_dict)
    return patient_level_split_assignments


def split_document_data(doc_dev_dict, doc_test_dict, doc_train_dict, assignments):
    full_data_set_dir = c.data_repo_dir
    for file in os.listdir(full_data_set_dir):
        p_d = file.split("_")
        patient_id = p_d[0]
        doc_id = p_d[1]

        try:
            assignment = assignments[patient_id]
            if assignment == "DEV":
                if not patient_id + "-" + doc_id in doc_dev_dict.keys():
                    doc_dev_dict[patient_id + "-" + doc_id] = "NO_LABEL"
            elif assignment == "TEST":
                if not patient_id + "-" + doc_id in doc_test_dict.keys():
                    doc_test_dict[patient_id + "-" + doc_id] = "NO_LABEL"
            elif assignment == "TRAIN":
                if not patient_id + "-" + doc_id in doc_train_dict.keys():
                    doc_train_dict[patient_id + "-" + doc_id] = "NO_LABEL"
        except KeyError:
            print "Error: Key " + patient_id + " does not have GOLD or SILVER labels! skipping..."
    write_new_splits_to_file("documents", doc_dev_dict, doc_test_dict, doc_train_dict)


def get_splits():
    # load silver data for easy annotation mapping
    mrn_caisis_dict, caisis_silver_dict = load_caisis_silver_annotations()

    doc_dev_gold_dir = c.doc_dev_gold_dir
    doc_test_gold_dir = c.doc_test_gold_dir
    doc_train_gold_dir = c.doc_train_gold_dir
    patients_test_gold_dir = c.patients_test_gold_dir
    patients_train_gold_dir = c.patients_train_gold_dir
    patients_dev_gold_dir = c.patients_dev_gold_dir

    dev_dirs = [doc_dev_gold_dir, patients_dev_gold_dir]
    test_dirs = [doc_test_gold_dir, patients_test_gold_dir]
    train_dirs = [doc_train_gold_dir, patients_train_gold_dir]

    dirs = [dev_dirs, test_dirs, train_dirs]

    doc_test_dict = dict()
    doc_dev_dict = dict()
    doc_train_dict = dict()

    patients_test_dict = dict()
    patients_dev_dict = dict()
    patients_train_dict = dict()

    def populate_dir_dict(doc_dict, patients_dict, dirs):
        doc_gold_dir = dirs[0]
        patients_gold_dir = dirs[1]

        with open(doc_gold_dir) as d:
            lines = d.readlines()
        for line in lines:
            id_label = line.split()
            id = id_label[0]
            label = id_label[1]
            doc_dict[id] = label

        with open(patients_gold_dir) as d:
            lines = d.readlines()
        for line in lines:
            id_label = line.split()
            id = id_label[0]
            label = id_label[1]
            patients_dict[id] = label
        pass

    populate_dir_dict(doc_dev_dict, patients_dev_dict, dev_dirs)
    populate_dir_dict(doc_test_dict, patients_test_dict, test_dirs)
    populate_dir_dict(doc_train_dict, patients_train_dict, train_dirs)

    # split and write patient data into dev/test/train, using Flor's gold labels and falling back to caisis silver
    #  when necessary
    patient_level_split_assignments = split_patient_data(patients_dev_dict, patients_test_dict, patients_train_dict,
                                                         caisis_silver_dict)

    # split and write document data into dev/test/train, using Flor's gold labels, and falling back to NO_LABEL
    #  when necessary
    split_document_data(doc_dev_dict, doc_test_dict, doc_train_dict, patient_level_split_assignments)

    pause = 9


def main():
    get_splits()


if __name__ == '__main__':
    main()
