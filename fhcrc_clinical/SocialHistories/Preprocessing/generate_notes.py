import openpyxl
from os import listdir
from os.path import isfile, join
import SystemUtilities.Configuration as c
import SystemUtilities.Globals as g
import re

from Preprocessing.PreDoc import PreDoc


def main():
    blobs = get_blobs()  # returns blobs in chronological order
    documents = get_docs_from_blobs(blobs)
    documents = add_caisis_ids_to_documents(documents)
    patients = assign_docs_to_patients(documents)
    flors_files = load_florText_florDocNum_dict()
    write_note_files_to_disk(patients, flors_files)


def add_caisis_ids_to_documents(documents):
    # load silver data for easy annotation mapping
    mrn_caisis_dict, caisis_silver_dict = load_caisis_silver_annotations()
    for doc in documents:
        if doc.mrn in mrn_caisis_dict:
            doc.caisis_id = mrn_caisis_dict[doc.mrn]
    return documents


def remove_date_dashes(event_date):
    event_date = event_date.replace("-", "")
    event_date = event_date.replace(":", "")
    event_date = event_date.replace(" ", "")
    return event_date.split()[
        0]  # theres an hour stamp im cutting off by doing this. no theres not. i need those hour nums


def get_blobs():
    # read blob file into memory
    with open(c.RAW_DATA_DIR, "rb") as f:
        data = f.read()

    # split on the defined delimiter
    blobs = re.split(g.DELIMITER, data)

    # Order blob by timestamps
    blobs2 = list()  # blobs2 holds a list of Tuple(date_order_num, blob)
    for blob in blobs:
        # split Metadata from the text. \t delimited.
        metadata_and_text = re.split(r"\t", blob)
        if len(metadata_and_text) > g.BLOB_FIELDS:
            event_date = metadata_and_text[2]
            date_order_num = remove_date_dashes(event_date)
            blobs2.append((date_order_num, blob))

    # sort blobs2 by date_order_num
    blobs2.sort()
    return blobs2


def populate_split_dict(doc_dir, patients_dir):
    the_dict=dict()
    with open(doc_dir) as d:
        lines = d.readlines()
    for line in lines:
        id_label = line.split()
        id = id_label[0]
        label = id_label[1]
        the_dict[id] = label
    with open(patients_dir) as d:
        lines = d.readlines()
    for line in lines:
        id_label = line.split()
        id = id_label[0]
        label = id_label[1]
        the_dict[id] = label
    return the_dict


def populate_dir_dict():
    dev_dict= populate_split_dict(c.doc_dev_gold_dir, c.patients_dev_gold_dir)
    train_dict =populate_split_dict(c.doc_train_gold_dir, c.patients_train_gold_dir)
    test_dict = populate_split_dict(c.doc_test_gold_dir, c.patients_test_gold_dir)
    return dev_dict,train_dict,test_dict


def write_note_files_to_disk(patients, flors_files):
    document_metadata= dict()
    dev_dict, train_dict, test_dict = populate_dir_dict()
    dev_docs = list()
    train_docs = list()
    test_docs = list()
    unknown = list()
    for patient_id in patients.keys():
        doc_num_count = 0
        alt_num_count = 0
        documents = patients[patient_id]

        if patient_id == None:
            patient_id = "3744"

        # Divvy up the docs into flor's scheme vs our scheme
        this_data_is_from_flor = list()
        data_unique_to_our_set = list()
        for doc in documents:
            normalized_note_text = normalize_note_text(doc.text)
            if normalized_note_text in flors_files:
                obscured_id = flors_files[normalized_note_text].replace("-", "_")
                this_data_is_from_flor.append((obscured_id, doc.text))
            else:
                obscured_id = str(patient_id) + "_" + str(alt_num_count)
                data_unique_to_our_set.append((obscured_id, doc.text))
                alt_num_count += 1

            # record metadata for each doc, will be written to METADATA_OUTPUT dir
            document_metadata[obscured_id] = (doc.mrn, doc.timestamp)

        # write files
        for tup in this_data_is_from_flor:
            with open(c.NOTE_OUTPUT_DIR + tup[0], "w") as writefile:
                for line in re.split(r"\n", tup[1]):
                    writefile.write(line + "\n")
        for tup in data_unique_to_our_set:
            with open(c.NOTE_OUTPUT_DIR + tup[0], "w") as writefile:
                for line in re.split(r"\n", tup[1]):
                    writefile.write(line + "\n")
    print("Raw data written into individual document txt files at: " + c.NOTE_OUTPUT_DIR)

    # write metadata file
    with open(c.SUBSTANCE_IE_DATA_FOLDER + "marvelously_massive_metadata_muniments_dict.txt", "wb") as file:
        for key_id, value_tuple in document_metadata.iteritems():
            file.write(key_id + "\t" + value_tuple[0] + "\t" + value_tuple[1] + "\n")
    print("Metatata for keyword-filtered documents written to: " + c.SUBSTANCE_IE_DATA_FOLDER + "marvelously_massive_metadata_muniments_dict.txt")
    pass

def get_docs_from_blobs(blobs):
    documents = list()
    for date_num, blob in blobs:
        # split Metadata from the text. \t delimited.
        metadata_and_text = re.split(r"\t", blob)
        if len(metadata_and_text) > g.BLOB_FIELDS:
            parent_id = metadata_and_text[0]
            clinical_event_id = metadata_and_text[1]
            event_date = metadata_and_text[2]
            event_description = metadata_and_text[3]
            addtnl_event_description = metadata_and_text[4]
            doc_text = ""
            # Text sometimes spans more than one \t-delimited partition. Take it all, except the last partition, which is the MRN
            for idx in range(8, len(metadata_and_text) - 3, 1):
                doc_text += metadata_and_text[idx].lstrip().rstrip() + "\n"
            mrn = metadata_and_text[len(metadata_and_text) - 2]

            # create document
            doc = PreDoc(event_date, doc_text, mrn)
            documents.append(doc)
    return documents


def load_caisis_silver_annotations():
    # Read in metadata file (excel), creating:
    #  dict of {doc_id:patient_id}
    #  dict of {patient_id : [gold labels] }
    wb = openpyxl.load_workbook(c.CAISIS_DIR)
    sheets = wb.get_sheet_names()
    mrn_caisis_dict = dict()
    caisis_gold_dict = dict()
    sheet_name = sheets[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    for i in range(1, g.CAISIS_ROWS, 1):  # the num of rows in metadata file
        caisis_id = str(sheet.cell(row=i, column=2).value)
        mrn = str(sheet.cell(row=i, column=1).value)
        mrn_caisis_dict[mrn] = caisis_id
        if caisis_id not in caisis_gold_dict:
            caisis_gold_dict[caisis_id] = list()
        tobac_label = sheet.cell(row=i, column=3).value
        alc_label = sheet.cell(row=i, column=8).value
        caisis_gold_dict[caisis_id].append(alc_label)
        caisis_gold_dict[caisis_id].append(tobac_label)
    return mrn_caisis_dict, caisis_gold_dict


def assign_docs_to_patients(documents):
    patients = dict()
    for doc in documents:
        if doc.caisis_id in patients:
            patients[doc.caisis_id].append(doc)
        else:
            patients[doc.caisis_id] = list()
            patients[doc.caisis_id].append(doc)
    return patients


def insert_appropriate_num_of_zeros(doc_num_count):
    TOTAL_NUMS = 5
    leng = len(str(doc_num_count))
    final_str = ""
    for x in range(0, TOTAL_NUMS - leng, 1):
        final_str += "0"
    return final_str


def load_florText_florDocNum_dict():
    id_note = dict()
    onlyfiles = [f for f in listdir(c.FLORIAN_FULL_DATA) if isfile(join(c.FLORIAN_FULL_DATA, f))]
    for file in onlyfiles:
        with open(c.FLORIAN_FULL_DATA + file) as f:
            text = f.read()
        normalized_note_text = normalize_note_text(text)
        if normalized_note_text != "":
            id_note[normalized_note_text] = file
    return id_note


def normalize_note_text(text):
    text = re.sub("\s+", "", text)
    return text


if __name__ == '__main__':
    main()
