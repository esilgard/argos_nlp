import os

from fhcrc_clinical.SocialHistories.DataLoading import ServerQuery
from fhcrc_clinical.SocialHistories.DataModeling.DataModels import Span, AnnotatedAttribute
from fhcrc_clinical.SocialHistories.SystemUtilities.Configuration import DATA_DIR
from fhcrc_clinical.SocialHistories.SystemUtilities.Globals import ATTRIBS
import openpyxl
import cPickle as Pickle


def main():
    """ Create patients objects where values are those found in the adjudicated excel file and the spans are the
     union of spans across the annotators"""
    annotations = ServerQuery.get_annotations_from_server()

    # create adjudicated 'patients' object -- List[Patient]
    values = get_adjudicated_values()
    spans = {}

    annotators = annotations.keys()
    last_annotator = annotators[-1]
    patients = annotations[last_annotator]

    # Keep track of all but last annotator's spans
    for annotator in annotations:
        if annotator != last_annotator:
            for mrn, docs in annotations[annotator].items():
                for doc_id, events in docs.items():
                    update_doc_spans(doc_id, patients, mrn, spans, events)

    # Add adjudicated value and unioned spans
    for mrn, docs in annotations[last_annotator].items():
        for doc_id, events in docs.items():
            if doc_id not in patients[mrn]:
                spans[doc_id] = {}

            for substance, event in events.items():
                substance = event.substance_type

                for attribute in ATTRIBS[substance]:
                    field_name = event.substance_type + attribute

                    # Value
                    add_value(doc_id, values, field_name, event)

                    # Spans
                    if attribute in event.attributes:
                        add_spans(spans[doc_id], field_name, event.attributes[attribute].all_value_spans)

                    if field_name in spans[doc_id]:
                        all_spans_of_field = determine_all_spans(spans[doc_id][field_name])
                        if all_spans_of_field:
                            if attribute not in event.attributes:
                                event.attributes[attribute] = AnnotatedAttribute(field_name, all_spans_of_field, '')

                            event.attributes[attribute].all_value_spans = determine_all_spans(spans[doc_id][field_name])

    # save to pickle
    Pickle.dump(patients, open(DATA_DIR + "adjudicated_batch_o", "wb"))


def get_adjudicated_values():
    """
    Return values from excel file containing adjudicated batch 0 values
    :return: {doc_id: {field_name: value}}
    """
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "Adjudication"))
    sheets = wb.get_sheet_names()
    sheet_name = sheets[0]
    sheet = wb.get_sheet_by_name(sheet_name)

    num_of_rows = sheet.max_row
    doc_column = 1
    field_column = 2
    value_column = 7

    values = {}  # {doc_id: {field_name: value}}

    for i in range(1, num_of_rows):  # the num of rows in metadata file
        # Read info from file
        doc = str(sheet.cell(row=i, column=doc_column).value)
        field = str(sheet.cell(row=i, column=field_column).value)
        value = str(sheet.cell(row=i, column=value_column).value)

        # Add to dictionary
        if doc not in values:
            values[doc] = {}

        values[doc][field] = value

    return values


def add_value(doc_id, values, field_name, event):
    if doc_id in values:
        if field_name in values[doc_id]:
            event.status = values[doc_id][field_name]
    else:
        print("Why u no here " + doc_id + "?")


def update_doc_spans(doc_id, patients, mrn, spans, events):
    # Add doc if not previously seen
    if doc_id not in spans:
        spans[doc_id] = {}

    # Find spans for each attribute of each substance
    for substance, event in events.items():
        for attribute in event.attributes:
            field_name = substance + attribute

            '''
            # Add field for doc if not previously seen
            if field_name not in spans[doc_id]:
                spans[doc_id][field_name] = []
            '''

            # Spans
            add_spans(spans[doc_id], field_name, event.attributes[attribute].all_value_spans)


def add_spans(doc_spans, field_name, new_spans):
    # Get tracking of characters covered by spans thus far
    characters_covered = get_characters_covered(field_name, doc_spans, new_spans)

    # Update characters covered by current annotator
    for span in new_spans:
        for index in range(span.start, span.stop):
            characters_covered[index] = True


def get_characters_covered(field_name, doc_spans, new_spans):
    """
    :return: {character index: True/False}, where True indicated the span from index to index+1 is covered by some
    annotator's highlighted region.
    """
    # Initialize if not initialized
    if field_name not in doc_spans:
        end_of_last_span = new_spans[-1].stop
        doc_spans[field_name] = {i: False for i in range(end_of_last_span)}

    return doc_spans[field_name]


def determine_all_spans(doc_text_indexes_covered):
    """
    @type doc_text_indexes_covered: {character index: True/False}
    :return: [Span]
    """

    spans = []
    in_span = False
    span_start = 0
    span_end = 0

    for index, index_is_covered in doc_text_indexes_covered.items():
        if not in_span:
            if index_is_covered:
                # Not already in span and character is covered: Begin a new span
                span_start = index
                span_end = index + 1
                in_span = True

        else:
            # Already in a span
            if index_is_covered:
                # Keep it going
                span_end += 1
            else:
                # Span just ended
                spans.append(Span(span_start, span_end))
                in_span = False

    # Append any span ending at the end of the text
    if in_span:
        spans.append(Span(span_start, span_end))

    return spans


if __name__ == '__main__':
    main()
