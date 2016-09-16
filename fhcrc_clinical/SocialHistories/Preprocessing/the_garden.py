# coding=utf-8
'''
This script reads a file containing multiple blobs delineated by Emily's 1234567890qwertyuiop string and produces
a document with the usrId_documentId as the filename.
'''
import re
import openpyxl
import fhcrc_clinical.SocialHistories.Preprocessing.get_splits as annotation_splitter
import fhcrc_clinical.SocialHistories.SystemUtilities.Configuration as c
import fhcrc_clinical.SocialHistories.SystemUtilities.Globals as g
from os import listdir
from os.path import isfile, join

from fhcrc_clinical.SocialHistories.Preprocessing.PreDoc import PreDoc
from fhcrc_clinical.SocialHistories.Preprocessing.PrePatient import PrePatient


def load_caisis_silver_annotations():
    # Read in metadata file (excel), creating:
    #  dict of {doc_id:patient_id}
    #  dict of {patient_id : [gold labels] }
    wb = openpyxl.load_workbook(c.CAISIS_DIR)
    sheets = wb.get_sheet_names()
    mrn_caisis_dict = dict()
    caisis_gold_dict = dict()
    sheet_name = sheets[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    for i in range(1, g.CAISIS_ROWS, 1):  # the num of rows in metadata file
        caisis_id = str(sheet.cell(row=i, column=2).value)
        mrn = str(sheet.cell(row=i, column=1).value)
        mrn_caisis_dict[mrn] = caisis_id
        if caisis_id not in caisis_gold_dict.keys():
            caisis_gold_dict[caisis_id] = list()
        tobac_label = sheet.cell(row=i, column=3).value
        alc_label = sheet.cell(row=i, column=8).value
        caisis_gold_dict[caisis_id].append(alc_label)
        caisis_gold_dict[caisis_id].append(tobac_label)
    return mrn_caisis_dict, caisis_gold_dict


def get_special_doc_id(count, patient_id, doc_text, flor_data_dict, patient_doc_count,
                       insertion_doc_count, id_timestamp_text_dict):

    TOTAL_DIGITS = 5
    num_digits = len(str(count))
    num = ""
    for i in range(0, TOTAL_DIGITS - num_digits, 1):
        num += "0"
    final_num = num + str(count)

    if patient_id + "-" + final_num in flor_data_dict.keys():
        florsData = flor_data_dict[patient_id + "-" + final_num].split()
        our_data = doc_text.split()
        if florsData == our_data:
            return final_num

    # increment new document counter
    if patient_id in insertion_doc_count:
        insertion_doc_count[patient_id] += 1
    else:
        insertion_doc_count[patient_id] = 0

    patient_doc_count[patient_id] -= 1
    return num + "n" + str(insertion_doc_count[patient_id])


# First pass through blobs, extract date so we can sort by date
def remove_date_dashes(event_date):
    event_date = event_date.replace("-", "")
    event_date = event_date.replace(":", "")
    event_date = event_date.replace(" ", "")
    return event_date.split()[
        0]  # theres an hour stamp im cutting off by doing this. no theres not. i need those hour nums


def get_caisis_silver_labels(label_list):
    str = ""
    count = 0
    for item in label_list:
        if count == 0:  # Alcohol info
            str += "ALC="
        elif count == 1:  # Tobacco info
            str += "TOB="
        str += item + "##"
        count += 1
    return str.rstrip("##")


def write_doc_level_gold_file(pat_id, doc_id):
    pass


def load_flor_note_docs():
    id_note = dict()
    onlyfiles = [f for f in listdir(c.FLORIAN_FULL_DATA) if isfile(join(c.FLORIAN_FULL_DATA, f))]
    for file in onlyfiles:
        with open(c.FLORIAN_FULL_DATA + file) as f:
            text = f.read()
        id_note[file] = text
    return id_note


def load_florians_gold_labels():
    # populate dict for document-level annotations
    doc_label_florGold_dict = dict()
    with open(c.doc_all_gold_dir, "r") as gold_docs_file:
        lines = gold_docs_file.readlines()
    for line in lines:
        id_label = line.split()
        id = id_label[0]
        label = id_label[1]
        doc_label_florGold_dict[id] = label

    # populate dict for ptient-level annotations
    patient_label_florGold_dict = dict()
    with open(c.patients_all_gold_dir, "r") as gold_pats_file:
        lines = gold_pats_file.readlines()
    for line in lines:
        id_label = line.split()
        id = id_label[0]
        label = id_label[1]
        patient_label_florGold_dict[id] = label
    return patient_label_florGold_dict, doc_label_florGold_dict

def add_timestamp_to_dict(first_id, event_date, doc_text, dictionary):
    if first_id in dictionary:
        if event_date in dictionary[first_id]:
            dictionary[first_id][event_date].append(doc_text)
        else:
            dictionary[first_id][event_date] = list()
            dictionary[first_id][event_date].append(doc_text)
    else:
        dictionary[first_id] = dict()
        dictionary[first_id][event_date] = list()
        dictionary[first_id][event_date].append(doc_text)
    pass

# read blob file into memory
with open(c.RAW_DATA_DIR, "rb") as f:
    data = f.read()

# load silver data for easy annotation mapping
mrn_caisis_dict, caisis_silver_dict = load_caisis_silver_annotations()

# load Florian's Gold annotations into dictionaries
patient_label_florGold_dict, doc_label_florGold_dict = load_florians_gold_labels()

# load Florians actual data into memory (yeah...this script is super memory-intensive)
flor_docId_text_dict = load_flor_note_docs()

# split on the defined delimiter
blobs = re.split(g.DELIMITER, data)
i = 1
blobs2 = list()  # blobs2 holds a list of Tuple(date_order_num, blob)
for blob in blobs:
    # split Metadata from the text. \t delimited.
    metadata_and_text = re.split(r"\t", blob)
    if len(metadata_and_text) > g.BLOB_FIELDS:
        event_date = metadata_and_text[2]
        date_order_num = remove_date_dashes(event_date)
        blobs2.append((date_order_num, blob))

# sort blobs2 by date_order_num
blobs2.sort()

# Open files for writing annotations, process blobs one at a time
# with open(c.NOTE_OUTPUT_GOLD_LABELS_DIR+"patient_GOLD") as patient_gold_file:
#     with open(c.NOTE_OUTPUT_GOLD_LABELS_DIR+"documents_GOLD", "w") as doc_gold_file:
with open(c.NOTE_OUTPUT_GOLD_LABELS_DIR + "patients_SILVER", "w") as patient_silver_file:
    patient_doc_count = dict()
    insertion_doc_count = dict()
    id_timestamp_text_dict = dict()
    prepatient_docs = dict()
    patients_dict = dict()
    for date_num, blob in blobs2:
        # split Metadata from the text. \t delimited.
        metadata_and_text = re.split(r"\t", blob)

        if len(metadata_and_text) > g.BLOB_FIELDS:
            parent_id = metadata_and_text[0]
            clinical_event_id = metadata_and_text[1]
            event_date = metadata_and_text[2]
            event_description = metadata_and_text[3]
            addtnl_event_description = metadata_and_text[4]
            doc_text = ""
            # Text sometimes spans more than one \t-delimited partition. Take it all, except the last partition, which is the MRN
            for idx in range(8, len(metadata_and_text) - 3, 1):
                doc_text += metadata_and_text[idx].lstrip().rstrip() + "\n"
            mrn = metadata_and_text[len(metadata_and_text) - 2]

            # Attempting to imitate Florian's unique document id scheme by numbering in chrono order
            if mrn_caisis_dict.has_key(mrn):
                first_id = mrn_caisis_dict[mrn]
            else:
                first_id = mrn
            if "U" in first_id:
                first_id = "3744"

            if first_id in patient_doc_count:
                patient_doc_count[first_id] += 1
            else:
                patient_doc_count[first_id] = 0
            count = patient_doc_count[first_id]


            # if the id is/contains a 10 digit number AND the first_id is vaguely numeric and not crazy junk
            if re.match(r"\D*(\d{10})\D*", clinical_event_id):
                # use first_id and event_date to create a dictionary of dictionaries like this:
                #   {first_id : {date : [document]}}
                # add_timestamp_to_dict(first_id, event_date, doc_text, id_timestamp_text_dict)

                # add_patient_and_doc_to_dict()
                pre_doc = PreDoc(event_date, doc_text)
                pre_patient = PrePatient(first_id)
                if pre_patient in prepatient_docs:
                    prepatient_docs[pre_patient].append(pre_doc)
                else:
                    prepatient_docs[pre_patient]=list()
                    prepatient_docs[pre_patient].append(pre_doc)






                ############################
                ### write blob to a file ###
                ############################
                # second_id = get_special_doc_id(count, first_id, doc_text, flor_docId_text_dict, patient_doc_count,
                #                                                                                 insertion_doc_count,
                #                                                                                 id_timestamp_text_dict)
                # with open(c.NOTE_OUTPUT_DIR + first_id + "_" + second_id,
                #           "w") as writefile:  # was clinical_event_id, not count
                #     for line in re.split(r"\n", doc_text):
                #         writefile.write(line + "\n")
                # #
                # ##########################################
                # ### write silver annotations to a file ###
                # ##########################################
                # # Write the patient-level silver doc: Alc and Tob from the Caisis excel file
                # if (first_id in caisis_silver_dict):
                #     patient_silver_file.write(
                #         first_id + " " + get_caisis_silver_labels(caisis_silver_dict[first_id]) + "\n")
                # else:
                #     patient_silver_file.write(first_id + " UNKNOWN" + "\n")
                #
                #
                #     # Divvy up the data into testing/dev/training splits
                #     # annotation_splitter.get_splits()
